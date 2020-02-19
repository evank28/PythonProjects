"""
This script is written to take a list of mail recipients, outputted from the
NationBuilder CRM system, and to prepare this list for efficient mailing by
postal mailing. The *main goal* of this is to prevent multiple pieces of mail
from being sent to the same address -- this is costly and inefficient. Instead,
names at one address are combined in a way that is consistent with natual
language, and letters are address appropriately to multiple recipients.
"""

import os

os.chdir("INSERT PATH")

import openpyxl

wb = openpyxl.load_workbook('INSERT FILE NAME')
wb2 = openpyxl.workbook()
sheetA = wb.get_sheet_by_name('Sheet')
sheetB = wb2.get_sheet_by_name('Sheet1')
print('Open worked')

households = {}
names = ""


def xstr(s):
    if s is None:
        return ''
    return str(s)


def xint(i):
    if i is None:
        return 0
    return int(i)


for row in range(2, sheetA.max_row + 1):
    fName = xstr(sheetA['A' + str(row)].value)
    lName = xstr(sheetA['B' + str(row)].value)
    a1 = xstr(sheetA['C' + str(row)].value)
    a2 = xstr(sheetA['D' + str(row)].value)
    city = xstr(sheetA['E' + str(row)].value)
    prov = xstr(sheetA['F' + str(row)].value)
    pCode = xstr(sheetA['G' + str(row)].value)
    aID = xstr(sheetA['H' + str(row)].value)

    if aID in households:
        if fName not in households[aID]["fNames"]:
            households[aID]["fNames"] += [fName]
        if {"fName": fName, "lName": lName} not in households[aID]["people"]:
            households[aID]["people"] += [{"fName": fName, "lName": lName}]
    else:

        households.setdefault(aID,
                              {"a1": a1, "a2": a2, "city": city, "prov": prov,
                               "pCode": pCode, "fNames": [fName],
                               "people": [{"fName": fName, "lName": lName}]})


def names_list_maker(list):
    non = len(list)
    finalString = ""
    counter = 1
    add = ""
    checkList = []
    for name in list:
        if name in checkList:
            continue
        if counter == 1:  # for first term
            add = name
        elif counter == non:  # for last term
            add = " & " + name
        else:  # for all middle terms
            add = ", " + name
        finalString += add
        counter += 1
        checkList += [name]
    return finalString


row = 2
for aID in households:
    sheetB['A' + str(row)] = aID
    sheetB['B' + str(row)] = households[aID]["a1"]
    sheetB['C' + str(row)] = households[aID]["a2"]
    sheetB['D' + str(row)] = households[aID]["city"]
    sheetB['E' + str(row)] = households[aID]["prov"]
    sheetB['F' + str(row)] = households[aID]["pCode"]
    namesList = households[aID]["fNames"]
    non = len(namesList)
    namesString = names_list_maker(namesList)
    sheetB['G' + str(row)] = namesString
    sheetB['H' + str(row)] = non
    people = households[aID]["people"]
    lNameList = []
    addressee = ""
    for person in people:
        if not (person["lName"] in lNameList):
            lNameList += [person["lName"]]
    inList = []
    if len(lNameList) == non:
        for person in people:
            inList += [person["fName"] + " " + person["lName"]]
        addressee = names_list_maker(inList)

    elif len(lNameList) == 1:
        if len(namesList) > 2:
            addressee = "The " + lNameList[0] + " Family"
        else:
            addressee = namesList[0] + " & " + namesList[1] + " " + lNameList[0]

    else:
        for person in people:
            inList += [person["fName"] + " " + person["lName"]]
        addressee = names_list_maker(inList)

    sheetB['I' + str(row)] = addressee
    print(addressee, str(len(lNameList)))
    row += 1

wb2.save('ResultHouseholds.xlsx')
print("There are " + str(row - 1) + " households and " + str(
    sheetA.max_row) + " names in this list.")
