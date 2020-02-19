import math
import operator
import openpyxl
import openpyxl.utils.cell as cell


def format_results(results_dict: dict) -> str:
    """
    Returns a multiline string to represent the numeric results of a tally that
    was stored and passed to the function as a dictionary.
    >>> format_results({'Mary Jane': 25, 'Alex Brown':52, 'John Smith':43})
    Mary Jane: 25
    Alex Brown: 52
    John Smith: 43
    """
    output = ""
    for option in results_dict:
        output += "{}: {}".format(option, results_dict[option]) + "\n"
    return output


# TODO: Add param number_of_winners so program can spit out multiple winners
def count_stv(ballots_table: str,
              details: bool):
    """
    Takes an excel file formatted as a Google Form Multiple Choice Grid and \
    returns the winner per the Single Transferable Vote algorithm, as well as
    the details if requested.

    >>> count_stv("ballots.xlsx", False)
    John Smith is the winnner.

    >>> count_stv("ballots2.xlsx", True)
    John Smith is the winner.
    ##########
    COUNT 1 RESULTS:
    Mary Jane: 25
    Alex Brown: 52
    John Smith: 43

    * Mary Jane was eliminated.

    ##########
    COUNT 2 RESULTS:
    Alex Brown: 59
    John Smith: 60
    Late Rejected: 1

    * John Smith is declared as the winner.
    """
    # read excel file
    wb = openpyxl.load_workbook(ballots_table)
    ws = wb.active
    headings = []
    options = {}

    # rename columns as needed for this particular file
    for col in range(2, ws.max_column + 1):
        value = ws[cell.get_column_letter(col) + str(1)].value
        name = value.split("[")[1:][0][:-1]
        if name is not None:
            headings += [name]
            options.setdefault(name, 0)

    # iterate through data, storing each ballot as a seperate dictionary entry
    ballots = {}
    for row in range(2, ws.max_row + 1):
        uid = ws['A' + str(row)].value
        if uid is not None:
            # each dictionary entry should have a dictionary (ballot) as its
            # value
            ballots.setdefault(uid, {})
            # each of these dictionaries should be formatted with integer keys
            # as the rank on the ballot, as long as the corresponding column was
            # filled out
            for col in range(2, ws.max_column + 1):
                rank = ws[cell.get_column_letter(col) + str(row)].value
                if col - 2 < len(headings) and rank != "NaN" and \
                        rank is not None:
                    ballots[uid][rank] = headings[col - 2]
    # tabulates any rejected ballots
    rejected = {k: v for k, v in ballots.items() if not v}
    # deletes any rejected ballots
    ballots = {k: v for k, v in ballots.items() if v}

    winner_found = False
    ballot_count = 1
    detailed_results = ""
    eliminated = []
    late_rejected_ballots = {}
    # loop through the ballots dictionary
    while not winner_found and len(options) >= 2:
        for uid in ballots:
            # iterate through the ballots dictionary,
            # calculating/recalculating the sum
            viable_ballot = False  # actually we don't know at this stage
            position = 1
            while not viable_ballot:
                # count the first place vote on each ballot and sum it in
                # another dictionary with all the values called options
                try:
                    choice = ballots[uid][position]
                except:
                    break
                # any time a ballot is encountered where the first place vote
                # has been eliminated, iterate down the ballot until an option
                # that has not been eliminated is available
                if not (choice in eliminated):
                    options[choice] += 1
                    viable_ballot = True
                else:
                    position += 1
            if not viable_ballot:
                # if all options on the ballot have been eliminated, remove the
                # ballot from the ballots dictionary and add it to the
                # late_rejected_ballots dictionary
                late_rejected_ballots.setdefault(uid, ballots.get(uid))
        for uid in late_rejected_ballots:
            try:
                del ballots[uid]
            except:
                pass  # print("could not remove " + str(uid))

        detailed_results += "\n##########\n" + "COUNT {} RESULTS:\n".format(
            ballot_count) + format_results(options)
        if ballot_count == 1:
            detailed_results += "Rejected: " + str(len(rejected)) + "\n"
        late_rejects = len(late_rejected_ballots)
        if late_rejects >= 1:
            detailed_results += "Late Rejected: " + str(late_rejects) + "\n"
        # if the top option has over 50% of the eligible votes, declare it the
        # winner and break from the loop
        if max(options.items(), key=operator.itemgetter(1))[1] >= (
                math.ceil(len(ballots) / 2) + 1):
            winner_found = True
            winner = max(options.items(), key=operator.itemgetter(1))[0]
            detailed_results += "\n* {} is declared as the winner.\n".format(
                winner)
        # otherwise, pop out the minimum option and add it's key to the list
        # eliminated. Continue looping through
        else:
            to_eliminate = min(options.items(), key=operator.itemgetter(1))[0]
            eliminated += [to_eliminate]
            try:
                del options[to_eliminate]
            except:
                pass
            detailed_results += "\n* {} was eliminated.\n".format(to_eliminate)
            ballot_count += 1
            for key in options:
                options[key] = 0

    if winner_found:
        if details:
            return winner + " is the winner.\n\n" + detailed_results
        return winner + " is the winner."
    # if the ballots dictionary is empty and no winner has been declared,
    # log an error
    else:
        return "ERROR: No winner found. \n" + detailed_results


if __name__ == "__main__":
    import os

    os.chdir("C:\\Users\\evank\\OneDrive\\Documents\\MyDevEnvironment\\"
             "My Python Projects\\Grad Clothing voting")
    print(count_stv("ballots_table.xlsx", True))
