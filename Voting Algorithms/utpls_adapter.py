#!/usr/bin/env python3
import csv
from typing import Dict, List, Set, Tuple

import openpyxl
import openpyxl.utils.cell as cell

from single_transferable_vote_new import count_stv_for_one_winner, Ballot, count_stv

FILE_NAME = 'UTPLS Voting Form 2021-2022 (Responses).xlsx'
UID_COL_LETTER = 'B'
ILLEGAL_VOTERS = set()

def get_ballots_for_election(election: str, workbook: openpyxl.workbook,
                             relevant_cols: List[str], allowed_uids: Set[str]) -> Tuple[Dict, Ballot]:
    ws = workbook.active
    # headings = []
    candidates = {}

    # # rename columns as needed for this particular file
    # for col in range(2, ws.max_column + 1):
    #     name = ws[cell.get_column_letter(col) + str(1)].value.strip()
    #     if name is not None and election.lower() in name.lower():
    #         # headings.append(name)
    #         # NOTE: This assumes the first column is first rank, second column is second rank,
    #         # and so on.
    #         relevant_cols.append(col)

    # iterate through data, storing each ballot as a separate dictionary entry
    ballots = {}
    for row in range(2, ws.max_row + 1):
        uid = str(int(ws[UID_COL_LETTER + str(row)].value))
        if uid is not None:
            if uid not in allowed_uids:
                global ILLEGAL_VOTERS
                ILLEGAL_VOTERS.add(uid)
                continue
            # each dictionary entry should have a Ballot
            # each of these dictionaries should be formatted with integer keys
            # as the rank on the ballot, as long as the corresponding column was
            # filled out
            cur_rank = 1
            rankings = {}
            for col in relevant_cols:
                value = ws[col + str(row)].value
                if not value:
                    continue
                candidate = value.strip()
                rankings[cur_rank] = candidate
                cur_rank += 1
                if candidate not in candidates and candidate != "Abstain":
                    candidates[candidate] = 0
            ballots[uid] = Ballot(uid, rankings)
    return ballots, candidates

#
# def is_rejection(cell_value: str) -> bool:
#     return cell_value == "Abstain"


if __name__ == "__main__":
    options_for_each_election = {
        "President": {"cols": ['C', 'D', 'E'], "num_winners": 1},
        "Vice President": {"cols": ['F', 'G'], "num_winners": 1},
        "Communications":  {"cols": ['H', 'I', 'J', 'K', 'L', 'M'], "num_winners": 4},
        "Canadian Mock Trial": {"cols": ['N', 'O', 'P', 'Q'], "num_winners": 2},
        "American Mock Trial": {"cols": ['R', 'S'], "num_winners": 2},
        "Mooting": {"cols": ['T', 'U'], "num_winners": 2},
        "Events": {"cols": ['V', 'W', 'X', 'Y'], "num_winners": 2},
        "Finance": {"cols": ['Z'], "num_winners": 1},
        "Mentorship": {"cols": ['AB', 'AC', 'AD'], "num_winners": 2}
    }
    with open('utpls_members.csv', 'r') as f:
        members = set(map(str.strip, f.readlines()))

    wb = openpyxl.load_workbook(FILE_NAME)
    data_for_each_election = \
        {election: get_ballots_for_election(election, wb,
                                            options_for_each_election[election]["cols"], members)
         for election in options_for_each_election.keys()}

    # results = count_stv_for_one_winner(*data_for_each_election['President'], details=True)
    # print(results)

    print(f"The following {len(ILLEGAL_VOTERS)} voters attempted to vote "
          f"illegally: {', '.join(ILLEGAL_VOTERS)}")

    for election in options_for_each_election.keys():
        winners, results_string = count_stv(*data_for_each_election[election], details=True,
                            number_of_winners=options_for_each_election[election]['num_winners'])
        with open(f"./results/{''.join(election.lower().split())}_results.txt", 'w') as f:
            print(f"# Results for Position: {election}", file=f)
            print(f"Number of winners: {options_for_each_election[election]['num_winners']}",
                  file=f)
            print(f"Number of Ballots Cast: {len(data_for_each_election[election][0])}", file=f)
            print(f"Winners: {', '.join(winners)}\n", file=f)
            print(f"Detailed results following:\n", file=f)
            print(results_string, file=f)


